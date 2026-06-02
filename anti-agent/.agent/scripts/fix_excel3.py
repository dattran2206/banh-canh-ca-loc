import json
import openpyxl
import re
import os

RESULTS_FILE = 'test-results/test-results.json'
EXCEL_FILE = '[VENTI]_Admin_Order Management_Testcase_VN.xlsx'

SHEET_MAPPING = {
    'ADM5': 'ADM5 Order - Subscription',
    'DETAIL': 'Order - Subscription Details'
}

def extract_specs(suite, specs_list):
    for spec in suite.get('specs', []):
        specs_list.append(spec)
    for sub_suite in suite.get('suites', []):
        extract_specs(sub_suite, specs_list)

def sync_results():
    if not os.path.exists(RESULTS_FILE):
        print(f"Error: {RESULTS_FILE} not found.")
        return

    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    test_outcomes = {}
    all_specs = []
    for suite in data.get('suites', []):
        extract_specs(suite, all_specs)

    for spec in all_specs:
        title = spec.get('title', '')
        match = re.search(r'\[(.*?)\]', title)
        if match:
            tag = match.group(1)
            tests = spec.get('tests', [])
            if tests:
                results = tests[0].get('results', [])
                if results:
                    status = results[0].get('status')
                    test_outcomes[tag] = 'Pass' if status in ['expected', 'passed'] else 'Fail'

    wb = openpyxl.load_workbook(EXCEL_FILE)
    updates = 0

    # Update ADM5
    ws_adm5 = wb[SHEET_MAPPING['ADM5']]
    adm5_counter = 1
    for row in range(14, ws_adm5.max_row + 1):
        steps = ws_adm5.cell(row=row, column=5).value
        expected = ws_adm5.cell(row=row, column=6).value
        if (steps and str(steps).strip()) or (expected and str(expected).strip()):
            tag = f"ADM5_{adm5_counter}"
            outcome = test_outcomes.get(tag)
            if outcome:
                # Column J is index 10 (1-based)
                ws_adm5.cell(row=row, column=10, value=outcome)
                updates += 1
            adm5_counter += 1

    # Update DETAIL
    ws_detail = wb[SHEET_MAPPING['DETAIL']]
    detail_counter = 1
    for row in range(14, ws_detail.max_row + 1):
        steps = ws_detail.cell(row=row, column=5).value
        expected = ws_detail.cell(row=row, column=6).value
        if (steps and str(steps).strip()) or (expected and str(expected).strip()):
            tag = f"DETAIL_{detail_counter}"
            outcome = test_outcomes.get(tag)
            if outcome:
                ws_detail.cell(row=row, column=10, value=outcome)
                updates += 1
            detail_counter += 1

    print(f"Total updates made: {updates}")
    wb.save(EXCEL_FILE)
    print("Excel file successfully updated with statuses.")

if __name__ == '__main__':
    sync_results()
