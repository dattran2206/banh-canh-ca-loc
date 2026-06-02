import json
import openpyxl
import re
import os

RESULTS_FILE = 'test-results/test-results.json'
EXCEL_FILE = '[VENTI]_Admin_Order Management_Testcase_VN.xlsx'

# Mapping sheet names
SHEET_MAPPING = {
    'ADM5': 'ADM5 Order - Subscription',
    'DETAIL': 'Order - Subscription Details'
}

def extract_specs(suite, specs_list):
    for spec in suite.get('specs', []):
        specs_list.append(spec)
    for sub_suite in suite.get('suites', []):
        extract_specs(sub_suite, specs_list)

def sync_results():
    if not os.path.exists(RESULTS_FILE):
        print(f"Error: {RESULTS_FILE} not found.")
        return

    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    test_outcomes = {}
    all_specs = []
    
    for suite in data.get('suites', []):
        extract_specs(suite, all_specs)

    for spec in all_specs:
        title = spec.get('title', '')
        
        # Extract tags like [ADM5_1] or [DETAIL_1]
        match = re.search(r'\[(.*?)\]', title)
        if match:
            tag = match.group(1)
            
            # Determine outcome
            tests = spec.get('tests', [])
            if tests:
                results = tests[0].get('results', [])
                if results:
                    status = results[0].get('status')
                    test_outcomes[tag] = 'Pass' if status in ['expected', 'passed'] else 'Fail'

    print(f"Parsed outcomes: {test_outcomes}")

    # Load Excel Workbook
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"Failed to load excel file: {e}")
        return

    # Update ADM5 Order - Subscription
    ws_adm5 = wb[SHEET_MAPPING['ADM5']]
    for row in ws_adm5.iter_rows(min_row=2, max_col=10):
        cell_id = row[0].value
        if cell_id and str(cell_id).startswith('ADM5_'):
            outcome = test_outcomes.get(str(cell_id))
            if outcome:
                row[9].value = outcome

    # Update Order - Subscription Details
    ws_detail = wb[SHEET_MAPPING['DETAIL']]
    for row in ws_detail.iter_rows(min_row=2, max_col=10):
        cell_id = row[0].value
        if cell_id and str(cell_id).isdigit():
            tag = f"DETAIL_{cell_id}"
            outcome = test_outcomes.get(tag)
            if outcome:
                row[9].value = outcome

    wb.save(EXCEL_FILE)
    print("Sync complete. Excel file updated.")

if __name__ == '__main__':
    sync_results()
