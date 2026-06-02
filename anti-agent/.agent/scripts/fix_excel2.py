import json
import openpyxl
import re
import os
import pandas as pd

RESULTS_FILE = 'test-results/test-results.json'
EXCEL_FILE = '[VENTI]_Admin_Order Management_Testcase_VN.xlsx'

SHEET_MAPPING = {
    'ADM5': 'ADM5 Order - Subscription',
    'DETAIL': 'Order - Subscription Details'
}

def extract_specs(suite, specs_list):
    for spec in suite.get('specs', []):
        specs_list.append(spec)
    for sub_suite in suite.get('suites', []):
        extract_specs(sub_suite, specs_list)

def sync_results():
    if not os.path.exists(RESULTS_FILE):
        print(f"Error: {RESULTS_FILE} not found.")
        return

    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    test_outcomes = {}
    all_specs = []
    for suite in data.get('suites', []):
        extract_specs(suite, all_specs)

    for spec in all_specs:
        title = spec.get('title', '')
        match = re.search(r'\[(.*?)\]', title)
        if match:
            tag = match.group(1)
            tests = spec.get('tests', [])
            if tests:
                results = tests[0].get('results', [])
                if results:
                    status = results[0].get('status')
                    test_outcomes[tag] = 'Pass' if status in ['expected', 'passed'] else 'Fail'

    print(f"Test outcomes to update: {test_outcomes}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    updates = 0

    # Read with pandas to get evaluated formulas
    xls = pd.ExcelFile(EXCEL_FILE)

    # Update ADM5
    df_adm5 = pd.read_excel(xls, sheet_name=SHEET_MAPPING['ADM5'], header=None)
    ws_adm5 = wb[SHEET_MAPPING['ADM5']]
    
    # Iterate pandas dataframe to find rows
    for idx, row in df_adm5.iterrows():
        cell_id = row[0]
        if pd.notnull(cell_id) and str(cell_id).startswith('ADM5_'):
            outcome = test_outcomes.get(str(cell_id))
            if outcome:
                # idx is 0-based. In Excel, row 1 is idx=0. So Excel row is idx + 1
                excel_row = idx + 1
                # Column J is index 10 (1-based)
                ws_adm5.cell(row=excel_row, column=10, value=outcome)
                updates += 1

    # Update DETAIL
    df_detail = pd.read_excel(xls, sheet_name=SHEET_MAPPING['DETAIL'], header=None)
    ws_detail = wb[SHEET_MAPPING['DETAIL']]
    
    for idx, row in df_detail.iterrows():
        cell_id = row[0]
        if pd.notnull(cell_id) and str(cell_id).isdigit():
            tag = f"DETAIL_{int(cell_id)}"
            outcome = test_outcomes.get(tag)
            if outcome:
                excel_row = idx + 1
                ws_detail.cell(row=excel_row, column=10, value=outcome)
                updates += 1

    print(f"Total updates made: {updates}")
    wb.save(EXCEL_FILE)
    print("Excel file successfully updated with statuses.")

if __name__ == '__main__':
    sync_results()
