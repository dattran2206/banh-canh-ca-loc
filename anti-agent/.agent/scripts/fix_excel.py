import json
import openpyxl
import re
import os

RESULTS_FILE = 'test-results/test-results.json'
EXCEL_FILE = '[VENTI]_Admin_Order Management_Testcase_VN.xlsx'

SHEET_MAPPING = {
    'ADM5': 'ADM5 Order - Subscription',
    'DETAIL': 'Order - Subscription Details'
}

def extract_specs(suite, specs_list):
    for spec in suite.get('specs', []):
        specs_list.append(spec)
    for sub_suite in suite.get('suites', []):
        extract_specs(sub_suite, specs_list)

def sync_results():
    if not os.path.exists(RESULTS_FILE):
        print(f"Error: {RESULTS_FILE} not found.")
        return

    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    test_outcomes = {}
    all_specs = []
    for suite in data.get('suites', []):
        extract_specs(suite, all_specs)

    for spec in all_specs:
        title = spec.get('title', '')
        match = re.search(r'\[(.*?)\]', title)
        if match:
            tag = match.group(1)
            tests = spec.get('tests', [])
            if tests:
                results = tests[0].get('results', [])
                if results:
                    status = results[0].get('status')
                    test_outcomes[tag] = 'Pass' if status in ['expected', 'passed'] else 'Fail'

    print(f"Test outcomes to update: {test_outcomes}")

    try:
        # Load data-only to read formula results
        wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        # Load normal to preserve formulas when saving
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"Failed to load excel file: {e}")
        return

    # Helper function to find column index of 'Round 1' or default to column J (index 10)
    # Actually, column J is col index 10 in 1-based openpyxl, but let's hardcode J which is column 10
    
    # Update ADM5
    ws_data_adm5 = wb_data[SHEET_MAPPING['ADM5']]
    ws_adm5 = wb[SHEET_MAPPING['ADM5']]
    
    updates = 0
    for row_idx, row in enumerate(ws_data_adm5.iter_rows(min_row=2, max_col=10), start=2):
        cell_id = row[0].value
        if cell_id and str(cell_id).startswith('ADM5_'):
            outcome = test_outcomes.get(str(cell_id))
            if outcome:
                # Column J is index 10 (1-based), or row[9] in 0-based iter_rows. So we update ws_adm5.cell(row=row_idx, column=10)
                ws_adm5.cell(row=row_idx, column=10, value=outcome)
                updates += 1

    # Update DETAIL
    ws_data_detail = wb_data[SHEET_MAPPING['DETAIL']]
    ws_detail = wb[SHEET_MAPPING['DETAIL']]
    
    for row_idx, row in enumerate(ws_data_detail.iter_rows(min_row=2, max_col=10), start=2):
        cell_id = row[0].value
        if cell_id and str(cell_id).isdigit():
            tag = f"DETAIL_{cell_id}"
            outcome = test_outcomes.get(tag)
            if outcome:
                ws_detail.cell(row=row_idx, column=10, value=outcome)
                updates += 1

    print(f"Total updates made: {updates}")
    wb.save(EXCEL_FILE)
    print("Excel file successfully updated with statuses.")

if __name__ == '__main__':
    sync_results()
