import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = [
    {
        "STT": 1,
        "Tên Vấn Đề / Issue Name": "Arbitrary SQL Execution",
        "Thành Phần / Component": "Backend Services",
        "Loại Lỗ Hổng / Vulnerability Type": "SQL Injection / RCE",
        "Mức Độ / Severity": "CRITICAL",
        "Mô Tả Chi Tiết / Description": "Server nhận và thực thi trực tiếp chuỗi SQL thô từ phía Client mà không qua kiểm soát.",
        "Vị Trí Cụ Thể / Location": "staging/server.js:304, web/routes/updateValue.js:17",
        "Tác Động / Impact": "Chiếm quyền điều khiển Database, xóa dữ liệu hoặc leo thang đặc quyền.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Loại bỏ tham số req.body.sql, sử dụng Parameterized Queries hoặc ORM.",
        "Tham Chiếu / References": "OWASP A03:2021",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    },
    {
        "STT": 2,
        "Tên Vấn Đề / Issue Name": "Docker Escape / Host Root Access",
        "Thành Phần / Component": "Infrastructure",
        "Loại Lỗ Hổng / Vulnerability Type": "Privilege Escalation",
        "Mức Độ / Severity": "CRITICAL",
        "Mô Tả Chi Tiết / Description": "Mount trực tiếp Docker socket của máy host vào container web.",
        "Vị Trí Cụ Thể / Location": "docker-compose.yml:90",
        "Tác Động / Impact": "Nếu ứng dụng web bị hack, kẻ tấn công có thể kiểm soát toàn bộ máy chủ vật lý.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Gỡ bỏ mount socket, sử dụng Docker API Proxy (read-only) nếu cần.",
        "Tham Chiếu / References": "CVE-2019-5736",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    },
    {
        "STT": 3,
        "Tên Vấn Đề / Issue Name": "Missing Authentication",
        "Thành Phần / Component": "API Layer",
        "Loại Lỗ Hổng / Vulnerability Type": "Broken Access Control",
        "Mức Độ / Severity": "HIGH",
        "Mô Tả Chi Tiết / Description": "Không có bất kỳ cơ chế xác thực nào cho các endpoint nhạy cảm.",
        "Vị Trí Cụ Thể / Location": "web/app.js, staging/server.js",
        "Tác Động / Impact": "Ai cũng có thể truy cập, xem và sửa đổi dữ liệu stub/staging từ Internet.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Triển khai Middleware xác thực (API Key hoặc OAuth2).",
        "Tham Chiếu / References": "OWASP A01:2021",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    },
    {
        "STT": 4,
        "Tên Vấn Đề / Issue Name": "Hardcoded Credentials",
        "Thành Phần / Component": "Config / Env",
        "Loại Lỗ Hổng / Vulnerability Type": "Insecure Storage",
        "Mức Độ / Severity": "HIGH",
        "Mô Tả Chi Tiết / Description": "Mật khẩu root MySQL và các dịch vụ để mặc định 'aA@123456'.",
        "Vị Trí Cụ Thể / Location": "docker-compose.yml:12, 46, 68, 86",
        "Tác Động / Impact": "Dễ dàng bị tấn công brute-force hoặc truy cập trái phép vào DB.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Sử dụng Docker Secrets hoặc file .env và đặt pass mạnh.",
        "Tham Chiếu / References": "OWASP A07:2021",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    },
    {
        "STT": 5,
        "Tên Vấn Đề / Issue Name": "Sensitive Data Leakage",
        "Thành Phần / Component": "Database",
        "Loại Lỗ Hổng / Vulnerability Type": "Data Privacy Violation",
        "Mức Độ / Severity": "HIGH",
        "Mô Tả Chi Tiết / Description": "Dữ liệu clone từ hệ thống Production chứa PII nhưng không được che giấu.",
        "Vị Trí Cụ Thể / Location": "Toàn bộ Database",
        "Tác Động / Impact": "Lộ lọt thông tin cá nhân khách hàng, vi phạm quy định bảo mật.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Triển khai quy trình Data Masking sau khi clone.",
        "Tham Chiếu / References": "GDPR / NIST",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    },
    {
        "STT": 6,
        "Tên Vấn Đề / Issue Name": "Unprotected Database Exposure",
        "Thành Phần / Component": "Network",
        "Loại Lỗ Hổng / Vulnerability Type": "Insecure Configuration",
        "Mức Độ / Severity": "MEDIUM",
        "Mô Tả Chi Tiết / Description": "Cổng MySQL (3306) và phpMyAdmin (8080) mở trực tiếp ra internet.",
        "Vị Trí Cụ Thể / Location": "docker-compose.yml:51, 73",
        "Tác Động / Impact": "Tăng bề mặt tấn công cho bot quét cổng.",
        "Giải Pháp Đề Xuất / Recommended Fix": "Đóng các port ra internet, dùng SSH Tunnel.",
        "Tham Chiếu / References": "OWASP A05:2021",
        "Trạng Thái / Status": "Open",
        "Ghi Chú / Notes": ""
    }
]

df = pd.DataFrame(data)
output_file = 'docs/SECURITY_CODE_REVIEW_REPORT.xlsx'

# Create with openpyxl
df.to_excel(output_file, index=False, engine='openpyxl')

# Load to style
from openpyxl import load_workbook
wb = load_workbook(output_file)
ws = wb.active

# Header Style
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Adjust Column Widths
column_widths = {
    'A': 5, 'B': 25, 'C': 15, 'D': 20, 'E': 10, 'F': 45, 'G': 35, 'H': 30, 'I': 45, 'J': 15, 'K': 10, 'L': 20
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Style data cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border

wb.save(output_file)
print(f'Successfully created {output_file}')
